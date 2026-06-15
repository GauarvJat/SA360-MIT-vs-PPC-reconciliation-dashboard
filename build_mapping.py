from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Column Mapping"

header_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

headers = ["Standardized Field", "SA360 Platform Source Column", "PPC Source Column", "Lumina Source Column", "Notes"]
ws.append(headers)
for c in ws[1]:
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

rows = [
    ["Year", "Derived from 'Month' (date parsed)", "Derived from 'Year/Month'", "Year", "Extracted via pandas datetime parsing"],
    ["Month", "Month_Number (derived from 'Month')", "Derived from 'Year/Month'", "Not used at month level", "Platform & PPC aligned to Year/Month before aggregation"],
    ["Platform", "Hardcoded = 'SA360'", "'Data Source'", "Hardcoded = 'SA360'", "Standardized across all 3 sources"],
    ["Advertiser Name", "'Sub-manager'", "'ADVERTISER NAME'", "n/a", "Source for Brand & Market_Code extraction"],
    ["Advertiser Id", "'Sub-manager ID' (dashes removed)", "'ADVERTISER_ID_FINAL' (last 3 digits removed)", "n/a", "Cleaned to numeric ID format"],
    ["Brand", "Extracted via regex: MB~(.*?)_MK", "Short_Brand: extracted via regex MB~(.*?)_MK", "Split from 'Lumina MVB' (Market_Platform_Brand)", "Brand naming convention consistent across sources"],
    ["Market_Code", "Extracted via regex: MK~(.*?)(?:_|$)", "Extracted via regex: MK~(.*?)(?:_|$)", "Mapped via Country_code_df lookup on 'Market'", "Used as key for final merge"],
    ["Market / Market Name", "'Market_Name' (via Market_code_df lookup on Market_Code)", "'Market Name'", "'Market' (split from Lumina MVB, title case)", "Country reference files used for platform & Lumina separately"],
    ["Campaign / Placement", "'Account name' -> 'Campaign'; 'Campaign' -> 'Placement'", "'Campaign', 'Placement'", "n/a", "Renamed for consistency"],
    ["Currency", "'Currency code'", "'Currency'", "n/a", ""],
    ["Impressions", "'Impr.' (commas/quotes stripped)", "'Impressions'", "n/a (not aggregated for Lumina)", "Converted to int, NaN filled with 0"],
    ["Clicks", "'Clicks' (commas/quotes stripped)", "'Clicks'", "n/a", "Converted to int, NaN filled with 0"],
    ["Cost / Spend (Platform Currency)", "'Cost'", "'Spend Platform Currency'", "n/a", "Numeric, rounded to 2 decimals"],
    ["Spend GBP", "n/a", "'Spend GBP'", "'Lumina Spends GBP'", "Used for Lumina vs PPC comparison"],
    ["%_Spend Diff_Lumina/PPC", "Calculated field", "Calculated field", "Calculated field", "abs(Lumina - PPC Spend GBP) / PPC Spend GBP * 100"],
    ["Spend Diff_Lumina/PPC", "Calculated field", "Calculated field", "Calculated field", "abs(Spend GBP - Lumina Spends GBP)"],
    ["%_Spend Diff_Platform/PPC", "Calculated field", "Calculated field", "Calculated field", "abs(Cost - Spend Platform Currency) / Spend Platform Currency * 100"],
    ["Spend Diff_Platform/PPC", "Calculated field", "Calculated field", "Calculated field", "abs(Spend Platform Currency - Cost)"],
    ["Exclusions", "Rows where 'Sub-manager' contains 'VAX' (pharma/vaccine) removed", "n/a", "n/a", "Per business rule (step 9 of project plan)"],
]

for r in rows:
    ws.append(r)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
    for cell in row:
        cell.alignment = wrap
        cell.border = border

widths = [28, 38, 38, 32, 45]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

ws.freeze_panes = "A2"

# Second sheet: Filename / file identification keywords
ws2 = wb.create_sheet("File Identification")
ws2.append(["File Role", "Identification Keyword in Filename", "Loaded As", "Notes"])
for c in ws2[1]:
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

ws2_rows = [
    ["SA360 Platform Merged File", "'platform' in filename", "platform_df", "Output of Part 2-3 merge step"],
    ["PPC Export", "'ppc' in filename", "ppc_df", "Last row (subtotal) is dropped on load"],
    ["Lumina Export", "'lumina' in filename", "lumina_df", "MVB column split into Market/Platform/Brand"],
    ["Market Master (Platform)", "Market_Master_SA360 for platform.xlsx", "Market_code_df", "Maps Market_Code -> Market_Name for platform data"],
    ["Country Name/Code (Lumina)", "country_name_code for lumina.xlsx", "Country_code_df", "Maps Market -> Market_Code for Lumina data"],
]
for r in ws2_rows:
    ws2.append(r)

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=4):
    for cell in row:
        cell.alignment = wrap
        cell.border = border

widths2 = [28, 40, 18, 45]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
ws2.freeze_panes = "A2"

wb.save("/home/claude/repo/Documentation/Data_Mapping.xlsx")
print("done")
